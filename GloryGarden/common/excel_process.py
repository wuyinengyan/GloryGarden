import xlrd
# import shutil
# import xlutils  # AttributeError: module 'xlutils' has no attribute 'copy'
from xlutils.copy import copy
from openpyxl import load_workbook


# （弃用）excel追加写入（xls：批注会丢失）
def write_excel_xls_append(file_path, df):  # df:DataFrame
    book = xlrd.open_workbook(file_path, formatting_info=True)  # 保持Excel原格式；type:book
    workbook = copy(book)  # type:workbook
    worksheet = workbook.get_sheet(0)  # == workbook.sheets[0]；type:worksheet
    for i in range(0, len(df)):  # 行
        for j in range(13):  # 列
            worksheet.write(i + 2, j, str(df.iloc[i][j]))  # 追加数据，从3行开始写入
    workbook.save(file_path)  # 保存工作簿


# excel追加写入（xlsx：批注不会丢失）
def write_excel_xlsx_append(file_path, df):  # df:DataFrame
    wb = load_workbook(file_path)  # type:workbook
    ws = wb.active  # type:worksheet
    for i in range(0, len(df)):  # 行
        for j, c in enumerate("ABCDEFGHIJKLM"):  # 列
            ws[f"{c}{i + 3}"] = str(df.iloc[i][j])  # 追加数据，从3行开始写入
    wb.save(file_path)


# 生成check的反馈文件
def check_excel_xlsx_append(file_path, df):  # df:DataFrame
    wb = load_workbook(file_path)  # type:workbook
    ws = wb.active  # type:worksheet
    for i in range(0, len(df)):  # 行
        for j, c in enumerate("ABCDEFGHIJ"):  # 列
            ws[f"{c}{i + 2}"] = str(df.iloc[i][j])  # 追加数据，从2行开始写入
    wb.save(file_path)
